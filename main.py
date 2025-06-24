import os
from classes import *

from openpyxl import load_workbook



input_dir = os.getcwd()
project_data_dir = '{}\\{}'.format(input_dir, '_project')
project_base_data_file = '{}\\{}\\data.xlsx'.format(project_data_dir, '01_Base_data')
project_plane_data = '{}\\{}'.format(project_data_dir, '02_Plane_data')
project_terrain_data = '{}\\{}'.format(project_data_dir, '03_Terrain_data')
project_drawings = '{}\\{}'.format(project_data_dir, '04_Drawings')

# pobierz dane z excela i twórz obiekty ekranów
ekrany = [] # do zmiany na obiekt

wb=load_workbook(project_base_data_file)
ws = wb['data']

for col in range(1,ws.max_column):
    
    screen = AcousticScreen()

    screen.number                           = ws[1][col].value
    screen.description_name                 = ws[2][col].value
    screen.height                           = ws[3][col].value
    screen.type                             = ws[4][col].value
    screen.terain_milage_dalay              = ws[5][col].value 
    screen.first_axis_number                = ws[6][col].value
    screen.mileage_delay                    = ws[7][col].value
    screen.previous_screen                  = ws[8][col].value
    screen.next_screen                      = ws[9][col].value
    screen.start_higher_load_axes_number    = ws[10][col].value
    screen.end_higher_load_axes_number      = ws[11][col].value

    screen.doors_position = [ws[row][col].value for row in range(12,ws.max_row+1) if ws[row][col].value]

    ekrany.append(screen)

for screen in ekrany[2:3]:
    try:
        # pobierz dane o terenie
        ws = wb[screen.description_name]
        terrain_data = [[ws[row][5].value,ws[row][6].value] for row in range(2,ws.max_row+1) if ws[row][6].value]
        screen.get_terrain_data(terrain_data)

        # tu dodać i zamienić to niżej na tworzenie obiektu osi oraz od razu wszystkich obiektów przyległych - pustych

        # pobierz dane o lokalizacjach pali - może do zmiany na osie, bo są głównym obiektem
        piles_data = [[ws[row][2].value,ws[row][3].value] for row in range(2,ws.max_row+1) if ws[row][2].value]
        screen.get_piles_data(piles_data)

        # pobierz dane o typach słupów - może do zmiany żeby ustalało automatycznie
        poles_data = [ws[row][1].value for row in range(2,ws.max_row+1) if ws[row][1].value]
        screen.get_poles_data(poles_data)

    except:
        print(f'brak danych dla ekranu: {screen.description_name}')

rysuj = 2


ekrany[rysuj].make()
ekrany[rysuj].draw_profil()
pass
